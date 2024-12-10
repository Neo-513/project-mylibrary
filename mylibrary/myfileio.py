import csv
import json
import os
import openpyxl
import pickle
import xlrd
import xlwt

DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop", "").replace("\\", "/")


def read(path: str, encoding: str = "utf-8") -> str | list | dict:
	extension = os.path.splitext(path)[1]

	if extension == ".pkl":
		with open(path, mode="rb") as f:
			datas = pickle.load(f)
	elif extension == ".csv":
		with open(path, mode="r", encoding=encoding) as f:
			datas = [[str(r) for r in reader] for reader in csv.reader(f)]
	elif extension == ".json":
		with open(path, mode="r", encoding=encoding) as f:
			datas = json.load(f)
	elif extension == ".xlsx":
		workbook = openpyxl.load_workbook(path, read_only=True)
		datas = {worksheet.title: [
			[str(d.value) if d.value is not None else "" for d in data] for data in worksheet.rows
		] for worksheet in workbook}

		if len(workbook.sheetnames) == 1:  # 单个工作表
			worksheet = workbook.active  # 获取当前工作表
			datas = datas[worksheet.title]
		workbook.close()
	elif extension == ".xls":
		workbook = xlrd.open_workbook(path)  # 获取工作簿
		datas = {worksheet.name: [
			[str(d.value) if d.value is not None else "" for d in data] for data in worksheet.get_rows()
		] for worksheet in workbook}

		if len(workbook.sheet_names()) == 1:  # 单个工作表
			worksheet = workbook.sheet_by_index(0)  # 获取当前工作表
			datas = datas[worksheet.name]
	else:
		with open(path, mode="r", encoding=encoding) as file:
			datas = file.read()

	print(f"[FILE READ]    {path}")
	return datas


def write(path: str, datas: str | list | dict, encoding: str = "utf-8"):
	extension = os.path.splitext(path)[1]

	if extension == ".pkl":
		with open(path, mode="wb") as f:
			pickle.dump(datas, f)
	elif extension == ".csv":
		with open(path, mode="w", encoding=encoding, newline="") as f:
			csv.writer(f).writerows([[str(d) for d in data] for data in datas])
	elif extension == ".json":
		with open(path, mode="w", encoding=encoding) as f:
			json.dump(datas, f, ensure_ascii=False)
	elif extension == ".xlsx":
		workbook = openpyxl.Workbook(write_only=True)  # 创建工作簿（只写模式）
		worksheet = workbook.create_sheet()  # 创建工作表

		for data in datas:
			worksheet.append([str(d) for d in data])
		workbook.save(path)  # 保存工作簿
		workbook.close()  # 关闭工作簿
	elif extension == ".xls":
		workbook = xlwt.Workbook()  # 创建工作簿
		worksheet = workbook.add_sheet("Sheet")  # 创建工作表

		for i, data in enumerate(datas):
			for j, d in enumerate(data):
				worksheet.write(i, j, str(d))
		workbook.save(path)  # 保存工作簿
	else:
		with open(path, mode="w", encoding=encoding) as file:
			file.write(datas)
	print(f"[FILE WRITTEN]    {path}")
